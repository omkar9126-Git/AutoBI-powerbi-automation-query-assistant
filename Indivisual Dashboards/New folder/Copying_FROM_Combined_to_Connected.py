import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog

# 🔐 Check if a file is locked (e.g., open in Excel)
def is_file_locked(filepath):
    try:
        with open(filepath, 'a'):
            return False
    except OSError:
        return True

# 🗂️ Pick source file using file dialog
Tk().withdraw()
source_file = filedialog.askopenfilename(
    title="Select the source combined_output Excel file",
    filetypes=[("Excel files", "*.xlsx")]
)
if not source_file:
    print("❌ No source file selected. Exiting.")
    exit()

# 🗃️ Set destination file path (change this to your own desired path)
destination_file = r"D:\internship\FINAL INTERNSHIP PROJECT\Combined Copied Worksheet\FINAL COMBINED TABLE WORKSHEET.xlsx"  # ⬅️ UPDATE THIS

# ❌ Exit if destination file is open/locked
if os.path.exists(destination_file) and is_file_locked(destination_file):
    print("❌ Destination file is open in Excel. Please close it and run again.")
    exit()

# ✅ Load data from source
df = pd.read_excel(source_file, sheet_name="CombinedData")

# 📘 Load or create destination workbook
if os.path.exists(destination_file):
    wb = load_workbook(destination_file)
    if "CombinedData" in wb.sheetnames:
        wb.remove(wb["CombinedData"])
else:
    wb = Workbook()
    # Remove default sheet if empty
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

# ➕ Create new sheet
ws = wb.create_sheet("CombinedData")

# ✍️ Write headers
for col_idx, col_name in enumerate(df.columns, 1):
    ws.cell(row=1, column=col_idx, value=col_name)

# ✍️ Write data rows
for row_idx, row in enumerate(df.values, 2):
    for col_idx, value in enumerate(row, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 🎨 Apply Excel Table formatting
end_col = get_column_letter(len(df.columns))
end_row = df.shape[0] + 1
table_range = f"A1:{end_col}{end_row}"
table = Table(displayName="CombinedTable", ref=table_range)
style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
table.tableStyleInfo = style
ws.add_table(table)

# 💾 Save updated file
wb.save(destination_file)
print(f"✅ Copied contents to: {destination_file}")
