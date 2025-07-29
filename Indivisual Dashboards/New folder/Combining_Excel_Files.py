import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog
from datetime import datetime

# 📂 Select folder
Tk().withdraw()
folder_path = filedialog.askdirectory(title="Select the folder containing Excel files")
if not folder_path:
    print("❌ No folder selected. Exiting.")
    exit()

# 🎯 Target columns
target_columns = [
    "ROW ID",
    "Heat No.", "Material", "Actual Amount", "Amount", "Actual Quantity", "Quantity",
    "Unit", "Conversion_to_MT", "Price per MT", "Final Product Name", "Final Product Price",
    "Final Pro. Quantity", "Yield", "Final LM output", "Final Bloom Output",
    "Yield for Scrap to LM", "Yield for Scrap to Bloom", "LM to Bloom Yield",
    "Source File", "Sheet Name", "Sections in Heat", "Section"
]

# 🔍 Keywords to match for section detection
section_keywords = ["UHPF", "CON", "VOD", "LRF"]

# 🧾 Filter Excel files
source_files = [
    os.path.join(folder_path, f)
    for f in os.listdir(folder_path)
    if f.endswith('.xlsx') and not f.startswith(('~$', 'combined_output'))
]

# 🔗 Build map of source file → all sections found in its sheet names
file_sections_map = {}
for file_path in source_files:
    try:
        xls = pd.ExcelFile(file_path)
        sheet_sections = set()
        for sheet_name in xls.sheet_names:
            for keyword in section_keywords:
                if keyword in sheet_name.upper():
                    sheet_sections.add(keyword)
        file_sections_map[os.path.basename(file_path)] = ", ".join(sorted(sheet_sections))
    except Exception as e:
        file_sections_map[os.path.basename(file_path)] = ""

combined_data = []

# 🔁 Read sheets and combine
for file_path in source_files:
    try:
        xls = pd.ExcelFile(file_path)
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            available_cols = [col for col in df.columns if col in target_columns]
            df_filtered = df[available_cols].copy()

            # Add missing columns
            for col in target_columns:
                if col not in df_filtered.columns:
                    df_filtered[col] = 0 if any(
                        key in col for key in ["Amount", "Quantity", "Yield", "Price"]
                    ) else ""

            file_name = os.path.basename(file_path)
            df_filtered["Source File"] = file_name
            df_filtered["Sheet Name"] = sheet_name
            df_filtered["Sections in Heat"] = file_sections_map.get(file_name, "")

            # 🔍 Detect section from sheet name
            sheet_section = ""
            for keyword in section_keywords:
                if keyword in sheet_name.upper():
                    sheet_section = keyword
                    break
            df_filtered["Section"] = sheet_section

            # Keep consistent column order (except ROW ID)
            ordered_cols = [col for col in target_columns if col != "ROW ID"]
            df_filtered = df_filtered[ordered_cols]

            combined_data.append(df_filtered)

    except Exception as e:
        print(f"❌ Error processing {file_path}: {e}")

# 🧩 Combine all
final_df = pd.concat(combined_data, ignore_index=True)

# 🔢 Add row IDs
final_df.insert(0, "ROW ID", range(1, len(final_df) + 1))

# 🧠 Fill missing Heat No. from filename
def extract_heat_no(filename):
    return os.path.splitext(filename)[0]

final_df["Heat No."] = final_df.apply(
    lambda row: row["Heat No."] if pd.notna(row["Heat No."]) and str(row["Heat No."]).strip() != ""
    else extract_heat_no(row["Source File"]),
    axis=1
)

# ✍️ Write to Excel with formatting
wb = Workbook()
ws = wb.active
ws.title = "CombinedData"

# Write headers
for col_idx, col_name in enumerate(target_columns, start=1):
    ws.cell(row=1, column=col_idx, value=col_name)

# Write rows
for row_idx, row in enumerate(final_df.values, start=2):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Add Excel table
end_col = get_column_letter(len(target_columns))
end_row = final_df.shape[0] + 1
table_range = f"A1:{end_col}{end_row}"
table = Table(displayName="CombinedTable", ref=table_range)
style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
table.tableStyleInfo = style
ws.add_table(table)

# Save with timestamp
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
output_filename = f"combined_output_{timestamp}.xlsx"
Tk().withdraw()
folder_path2 = filedialog.askdirectory(title="Select the folder to store combined outputs")
if not folder_path:
    print("❌ No folder selected. Exiting.")
    exit()
output_path = os.path.join(folder_path2, output_filename)
wb.save(output_path)

print(f"✅ Combined Excel file saved as: {output_filename}")
